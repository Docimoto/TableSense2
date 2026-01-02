"""
Tests for Excel utilities.
"""

import pytest
from openpyxl import Workbook
from utils.excel_utils import (
    excel_col_to_num,
    num_to_excel_col,
    parse_excel_range,
    num_to_excel_range,
    get_sheet_dimensions,
    get_merged_regions,
)


def test_excel_col_to_num():
    """Test Excel column letter to number conversion."""
    assert excel_col_to_num("A") == 1
    assert excel_col_to_num("B") == 2
    assert excel_col_to_num("Z") == 26
    assert excel_col_to_num("AA") == 27
    assert excel_col_to_num("AB") == 28
    assert excel_col_to_num("AZ") == 52
    assert excel_col_to_num("BA") == 53


def test_num_to_excel_col():
    """Test number to Excel column letter conversion."""
    assert num_to_excel_col(1) == "A"
    assert num_to_excel_col(2) == "B"
    assert num_to_excel_col(26) == "Z"
    assert num_to_excel_col(27) == "AA"
    assert num_to_excel_col(28) == "AB"
    assert num_to_excel_col(52) == "AZ"
    assert num_to_excel_col(53) == "BA"


def test_parse_excel_range():
    """Test Excel range parsing."""
    assert parse_excel_range("C8:H18") == (3, 8, 8, 18)
    assert parse_excel_range("A1:AB14") == (1, 1, 28, 14)
    assert parse_excel_range("A1:Z1") == (1, 1, 26, 1)
    
    with pytest.raises(ValueError):
        parse_excel_range("invalid")


def test_num_to_excel_range():
    """Test numeric coordinates to Excel range conversion."""
    assert num_to_excel_range(3, 8, 8, 18) == "C8:H18"
    assert num_to_excel_range(1, 1, 28, 14) == "A1:AB14"
    assert num_to_excel_range(1, 1, 26, 1) == "A1:Z1"


def test_get_sheet_dimensions():
    """Test sheet dimension retrieval."""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Test"
    ws['B2'] = "Test2"
    ws['C10'] = "Test3"
    
    height, width = get_sheet_dimensions(wb, ws.title)
    assert height == 10
    assert width == 3


def test_get_merged_regions():
    """Test merged region detection."""
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('A1:B2')
    ws.merge_cells('C3:D3')
    
    merged = get_merged_regions(ws)
    assert len(merged) == 2
    assert (1, 1, 2, 2) in merged
    assert (3, 3, 3, 4) in merged

