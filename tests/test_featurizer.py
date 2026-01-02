"""
Tests for cell featurizer.
"""

import pytest
import numpy as np
from openpyxl import Workbook
from features.featurizer import CellFeaturizer, NUM_FEATURES


def test_featurizer_output_shape():
    """Test that featurizer produces correct output shape."""
    featurizer = CellFeaturizer()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Test"
    ws['B2'] = 123
    
    features, metadata = featurizer.featurize_sheet(wb, ws.title)
    
    assert features.shape[2] == NUM_FEATURES  # 43 features
    assert features.shape[0] == ws.max_row
    assert features.shape[1] == ws.max_column


def test_featurizer_has_data():
    """Test has_data feature extraction."""
    featurizer = CellFeaturizer()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Text"  # Has data
    ws['B1'] = 123  # Has data
    ws['C1'] = ""  # Empty string
    ws['D1'] = None  # None value
    
    features, _ = featurizer.featurize_sheet(wb, ws.title)
    
    # Check has_data (index 0)
    assert features[0, 0, 0] == 1.0  # A1 has data
    assert features[0, 1, 0] == 1.0  # B1 has data
    assert features[0, 2, 0] == 0.0  # C1 empty string
    assert features[0, 3, 0] == 0.0  # D1 None


def test_featurizer_has_formula():
    """Test has_formula feature extraction."""
    featurizer = CellFeaturizer()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "=SUM(1,2)"  # Formula
    ws['B1'] = 123  # Not a formula
    
    features, _ = featurizer.featurize_sheet(wb, ws.title)
    
    # Check has_formula (index 1)
    assert features[0, 0, 1] == 1.0  # A1 has formula
    assert features[0, 1, 1] == 0.0  # B1 no formula


def test_featurizer_is_visibly_empty():
    """Test is_visibly_empty feature extraction."""
    featurizer = CellFeaturizer()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Text"  # Not empty
    ws['B1'] = ""  # Empty string
    ws['C1'] = None  # None value
    ws.merge_cells('D1:E1')
    ws['D1'] = "Merged"  # Top-left
    
    features, _ = featurizer.featurize_sheet(wb, ws.title)
    
    # Check is_visibly_empty (index 3)
    assert features[0, 0, 3] == 0.0  # A1 not empty
    assert features[0, 1, 3] == 1.0  # B1 empty string
    assert features[0, 2, 3] == 1.0  # C1 None
    # E1 (non-top-left in merged range) should be visibly empty
    assert features[0, 4, 3] == 1.0  # E1 merged non-top-left


def test_featurizer_is_formatted():
    """Test is_formatted feature extraction."""
    featurizer = CellFeaturizer()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 123
    ws['A1'].number_format = '0.00'  # Formatted
    ws['B1'] = 456  # General format (default)
    
    features, _ = featurizer.featurize_sheet(wb, ws.title)
    
    # Check is_formatted (index 2)
    assert features[0, 0, 2] == 1.0  # A1 formatted
    assert features[0, 1, 2] == 0.0  # B1 General format


def test_featurizer_merged_cells():
    """Test merged cell handling."""
    featurizer = CellFeaturizer()
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('A1:B2')
    ws['A1'] = "Merged"
    
    features, _ = featurizer.featurize_sheet(wb, ws.title)
    
    # All cells in merged region should have merge indicators
    # Check merge indicators (features 4-7: merged_left, merged_right, merged_top, merged_bottom)
    merge_start_idx = 4  # has_data(1) + has_formula(1) + is_formatted(1) + is_visibly_empty(1) = 4
    
    # A1 (top-left) should not have merged_left or merged_top
    assert features[0, 0, merge_start_idx] == 0.0  # merged_left
    assert features[0, 0, merge_start_idx + 2] == 0.0  # merged_top
    
    # B1 (top-right) should have merged_left
    assert features[0, 1, merge_start_idx] == 1.0  # merged_left
    
    # A2 (bottom-left) should have merged_top
    assert features[1, 0, merge_start_idx + 2] == 1.0  # merged_top


def test_featurizer_color_features():
    """Test color feature extraction."""
    featurizer = CellFeaturizer()
    wb = Workbook()
    ws = wb.active
    
    from openpyxl.styles import Font, PatternFill
    from openpyxl.styles.colors import Color
    
    # Set font color to red
    ws['A1'] = "Red text"
    ws['A1'].font = Font(color=Color(rgb="FFFF0000"))  # Red
    
    # Set background color to blue
    ws['B1'] = "Blue background"
    ws['B1'].fill = PatternFill(fgColor=Color(rgb="FF0000FF"))  # Blue
    
    features, _ = featurizer.featurize_sheet(wb, ws.title)
    
    # Check font color (features 8-15)
    font_color_start = 8  # has_data(1) + has_formula(1) + is_formatted(1) + is_visibly_empty(1) + merge(4) = 8
    # Red should be bucket 3
    assert features[0, 0, font_color_start + 3] == 1.0  # Red
    
    # Check background color (features 16-23)
    bg_color_start = 8 + 8  # font_color_start + COLOR_BUCKET_SIZE
    # Blue should be bucket 6
    assert features[0, 1, bg_color_start + 6] == 1.0  # Blue


def test_featurizer_bold_flag():
    """Test bold flag extraction."""
    featurizer = CellFeaturizer()
    wb = Workbook()
    ws = wb.active
    
    from openpyxl.styles import Font
    
    ws['A1'] = "Bold"
    ws['A1'].font = Font(bold=True)
    
    ws['B1'] = "Not bold"
    
    features, _ = featurizer.featurize_sheet(wb, ws.title)
    
    # Bold flag is at index 32
    # has_data(1) + has_formula(1) + is_formatted(1) + is_visibly_empty(1) + merge(4) + font_color(8) + bg_color(8) + digit_prop(1) + length(4) + font_size(3) = 32
    bold_idx = 32
    assert features[0, 0, bold_idx] == 1.0  # Bold
    assert features[0, 1, bold_idx] == 0.0  # Not bold


def test_featurizer_sample_tables():
    """Integration test with sample_tables.xlsx to verify all features match expectations."""
    import os
    from pathlib import Path
    from utils.excel_utils import load_workbook
    
    # Get path to sample_tables.xlsx
    test_data_dir = Path(__file__).parent / "data"
    sample_file = test_data_dir / "sample_tables.xlsx"
    
    if not sample_file.exists():
        pytest.skip(f"Sample file not found: {sample_file}")
    
    featurizer = CellFeaturizer()
    wb = load_workbook(sample_file, data_only=False)
    
    # Test Sheet1 table1: Sheet1!$A$1:$D$7
    # Features: Bold Font A1:D1, Formulas D2:D7, Format A2:A7
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        features, _ = featurizer.featurize_sheet(wb, "Sheet1")
        
        # Check bold flag (index 32) for A1:D1 (rows 0, cols 0-3)
        bold_idx = 32
        assert features[0, 0, bold_idx] == 1.0  # A1 bold
        assert features[0, 1, bold_idx] == 1.0  # B1 bold
        assert features[0, 2, bold_idx] == 1.0  # C1 bold
        assert features[0, 3, bold_idx] == 1.0  # D1 bold
        
        # Check has_formula (index 1) for D2:D7 (rows 1-6, col 3)
        formula_idx = 1
        for row in range(1, 7):  # D2:D7
            assert features[row, 3, formula_idx] == 1.0  # D2-D7 have formulas
        
        # Check is_formatted (index 2) for A2:A7 (rows 1-6, col 0)
        formatted_idx = 2
        for row in range(1, 7):  # A2:A7
            assert features[row, 0, formatted_idx] == 1.0  # A2-A7 formatted
    
    # Test Sheet2 table1: Sheet2!$A$1:$D$9
    # Features: Background yellow A1:D1, Bold A1:D1, Font black A1:D1
    if "Sheet2" in wb.sheetnames:
        ws = wb["Sheet2"]
        features, _ = featurizer.featurize_sheet(wb, "Sheet2")
        
        # Check background color yellow (bucket 4) for A1:D1
        bg_color_start = 16  # After font color (8)
        yellow_bucket = 4
        for col in range(4):  # A1:D1
            assert features[0, col, bg_color_start + yellow_bucket] == 1.0
        
        # Check bold flag
        bold_idx = 32
        for col in range(4):  # A1:D1
            assert features[0, col, bold_idx] == 1.0
        
        # Check font color black (bucket 0) for A1:D1
        font_color_start = 8
        black_bucket = 0
        for col in range(4):  # A1:D1
            assert features[0, col, font_color_start + black_bucket] == 1.0
    
    # Test Sheet4 table1: Sheet4!$A$1:$A$2
    # Features: digit_proportion A2, has_comment A1, size_l A2, is_locked A2
    if "Sheet4" in wb.sheetnames:
        ws = wb["Sheet4"]
        features, _ = featurizer.featurize_sheet(wb, "Sheet4")
        
        # Check has_comment (index 42) for A1
        comment_idx = 42
        assert features[0, 0, comment_idx] == 1.0  # A1 has comment
        
        # Check is_locked (index 38) for A2
        locked_idx = 38
        assert features[1, 0, locked_idx] == 1.0  # A2 locked
        
        # Check length bucket l (index 2) for A2
        length_start = 25
        length_l_bucket = 2
        assert features[1, 0, length_start + length_l_bucket] == 1.0  # A2 size_l
        
        # Check digit_proportion (index 24) for A2
        digit_prop_idx = 24
        # Should be approximately 0.07692307692 (1 digit out of 13 non-whitespace chars)
        digit_prop = features[1, 0, digit_prop_idx]
        assert 0.07 < digit_prop < 0.08  # Approximately 1/13
    
    wb.close()

