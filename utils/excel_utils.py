"""
Excel file utilities for working with openpyxl workbooks and worksheets.

Provides helper functions for:
- Sheet dimension queries
- Cell iteration
- Merged region detection
- Format extraction
- Excel coordinate conversion (A1 notation <-> numeric indices)
"""

from typing import Tuple, List, Iterator, Optional
import warnings
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook as _openpyxl_load_workbook

def load_workbook(
    path: Path | str, *args, **kwargs
) -> Workbook:
    """
    Wrap `openpyxl.load_workbook` so warnings include the workbook path.
    """
    normalized_path = Path(path)
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        workbook = _openpyxl_load_workbook(str(normalized_path), *args, **kwargs)

    for warn in caught:
        warnings.showwarning(
            message=f"{normalized_path}: {warn.message}",
            category=warn.category,
            filename=str(normalized_path),
            lineno=getattr(warn, "lineno", 0),
            line=getattr(warn, "line", None),
        )

    return workbook

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell


def get_sheet_dimensions(workbook: Workbook, sheet_name: str) -> Tuple[int, int]:
    """
    Get the dimensions (height, width) of a worksheet.
    
    Args:
        workbook: openpyxl Workbook object
        sheet_name: Name of the sheet
        
    Returns:
        Tuple of (height, width) in cells (1-based)
    """
    sheet = workbook[sheet_name]
    # Get the maximum row and column that have data
    if sheet.max_row is None or sheet.max_column is None:
        return (0, 0)
    return (sheet.max_row, sheet.max_column)


def iterate_cells(sheet: Worksheet) -> Iterator[Tuple[int, int, Cell]]:
    """
    Iterate over all cells in a worksheet with their coordinates.
    
    Args:
        sheet: openpyxl Worksheet object
        
    Yields:
        Tuples of (row, col, cell) where row and col are 1-based indices
    """
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            yield (row_idx, col_idx, cell)


def get_merged_regions(sheet: Worksheet) -> List[Tuple[int, int, int, int]]:
    """
    Get all merged cell regions in a worksheet.
    
    Args:
        sheet: openpyxl Worksheet object
        
    Returns:
        List of tuples (row_top, col_left, row_bottom, col_right) in 1-based indices
    """
    merged_regions = []
    for merged_range in sheet.merged_cells.ranges:
        # merged_range is a CellRange object
        # min_row, min_col, max_row, max_col are 1-based
        merged_regions.append((
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col
        ))
    return merged_regions


def get_effective_cell_coords(
    sheet: Worksheet, 
    row: int, 
    col: int, 
    merged_regions: Optional[List[Tuple[int, int, int, int]]] = None
) -> Tuple[int, int]:
    """
    Get the effective (top-left) cell coordinates for a given cell.
    
    If the cell is part of a merged region, returns the top-left cell coordinates.
    Otherwise, returns the original cell coordinates.
    
    Args:
        sheet: openpyxl Worksheet object
        row: 1-based row index
        col: 1-based column index
        merged_regions: Optional pre-computed merged regions list. If None, will compute.
        
    Returns:
        Tuple of (row, col) for the effective (top-left) cell in 1-based indices
    """
    if merged_regions is None:
        merged_regions = get_merged_regions(sheet)
    
    # Check if this cell is part of any merged region
    for row_top, col_left, row_bottom, col_right in merged_regions:
        if row_top <= row <= row_bottom and col_left <= col <= col_right:
            # This cell is part of a merged region, return top-left cell
            return (row_top, col_left)
    
    # Not part of any merged region, return original coordinates
    return (row, col)


def get_cell_format(sheet: Worksheet, row: int, col: int) -> dict:
    """
    Get formatting information for a specific cell.
    
    For merged cells, returns formatting from the top-left cell of the merged region.
    This ensures consistent formatting across all cells in a merged region.
    
    Args:
        sheet: openpyxl Worksheet object
        row: 1-based row index
        col: 1-based column index
        
    Returns:
        Dictionary with format information:
        - 'fill': cell fill (background color)
        - 'font': cell font (color, size, bold, etc.)
        - 'border': cell borders
        - 'alignment': cell alignment
        - 'number_format': cell number format
        - 'protection': cell protection settings
    """
    # Get effective cell coordinates (top-left if merged, otherwise original)
    eff_row, eff_col = get_effective_cell_coords(sheet, row, col)
    
    # Get formatting from the effective cell
    cell = sheet.cell(row=eff_row, column=eff_col)
    return {
        'fill': cell.fill,
        'font': cell.font,
        'border': cell.border,
        'alignment': cell.alignment,
        'number_format': cell.number_format,
        'protection': cell.protection,
    }


def excel_col_to_num(col_letter: str) -> int:
    """
    Convert Excel column letter(s) to 1-based numeric index.
    
    Args:
        col_letter: Column letter(s) (e.g., "A", "B", "AA", "AB")
        
    Returns:
        1-based column index (A=1, B=2, ..., Z=26, AA=27, etc.)
        
    Examples:
        >>> excel_col_to_num("A")
        1
        >>> excel_col_to_num("Z")
        26
        >>> excel_col_to_num("AA")
        27
        >>> excel_col_to_num("AB")
        28
    """
    result = 0
    for char in col_letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result


def num_to_excel_col(num: int) -> str:
    """
    Convert 1-based numeric column index to Excel column letter(s).
    
    Args:
        num: 1-based column index
        
    Returns:
        Column letter(s) (e.g., "A", "B", "AA", "AB")
        
    Examples:
        >>> num_to_excel_col(1)
        'A'
        >>> num_to_excel_col(26)
        'Z'
        >>> num_to_excel_col(27)
        'AA'
        >>> num_to_excel_col(28)
        'AB'
    """
    result = ""
    num -= 1  # Convert to 0-based
    while num >= 0:
        result = chr(ord('A') + (num % 26)) + result
        num = num // 26 - 1
    return result


def parse_excel_range(range_str: str) -> Tuple[int, int, int, int]:
    """
    Parse Excel range notation to numeric coordinates.
    
    Args:
        range_str: Excel range string (e.g., "C8:H18", "A1:AB14")
        
    Returns:
        Tuple of (col_left, row_top, col_right, row_bottom) in 1-based indices
        
    Examples:
        >>> parse_excel_range("C8:H18")
        (3, 8, 8, 18)
        >>> parse_excel_range("A1:AB14")
        (1, 1, 28, 14)
    """
    if ':' not in range_str:
        raise ValueError(f"Invalid range format: {range_str}")
    
    parts = range_str.split(':')
    if len(parts) != 2:
        raise ValueError(f"Invalid range format: {range_str}")
    
    start_cell = parts[0].strip()
    end_cell = parts[1].strip()
    
    # Parse start cell (e.g., "C8")
    start_col_letters = ""
    start_row = ""
    for char in start_cell:
        if char.isalpha():
            start_col_letters += char
        elif char.isdigit():
            start_row += char
    
    # Parse end cell (e.g., "H18")
    end_col_letters = ""
    end_row = ""
    for char in end_cell:
        if char.isalpha():
            end_col_letters += char
        elif char.isdigit():
            end_row += char
    
    if not start_col_letters or not start_row or not end_col_letters or not end_row:
        raise ValueError(f"Invalid range format: {range_str}")
    
    col_left = excel_col_to_num(start_col_letters)
    row_top = int(start_row)
    col_right = excel_col_to_num(end_col_letters)
    row_bottom = int(end_row)
    
    return (col_left, row_top, col_right, row_bottom)


def num_to_excel_range(col_left: int, row_top: int, col_right: int, row_bottom: int) -> str:
    """
    Convert numeric coordinates to Excel range notation.
    
    Args:
        col_left: 1-based left column index
        row_top: 1-based top row index
        col_right: 1-based right column index
        row_bottom: 1-based bottom row index
        
    Returns:
        Excel range string (e.g., "C8:H18")
        
    Examples:
        >>> num_to_excel_range(3, 8, 8, 18)
        'C8:H18'
        >>> num_to_excel_range(1, 1, 28, 14)
        'A1:AB14'
    """
    start_col = num_to_excel_col(col_left)
    end_col = num_to_excel_col(col_right)
    return f"{start_col}{row_top}:{end_col}{row_bottom}"


def get_excel_tables(sheet: Worksheet) -> dict:
    """
    Get Excel table objects from a worksheet.
    
    Args:
        sheet: openpyxl Worksheet object
        
    Returns:
        Dictionary mapping table names to (col_left, row_top, col_right, row_bottom) tuples
        in 1-based indices. Returns empty dict if no tables found or tables API unavailable.
    """
    excel_tables = {}
    
    if not hasattr(sheet, 'tables') or not sheet.tables:
        return excel_tables
    
    # sheet.tables is a TableList that can be iterated by keys (table names)
    # Access tables by name, not via .items() which returns strings
    for table_name in sheet.tables:
        table = sheet.tables[table_name]
        # table.ref is a string like "C14:F19"
        try:
            col_left, row_top, col_right, row_bottom = parse_excel_range(table.ref)
            excel_tables[table_name] = (col_left, row_top, col_right, row_bottom)
        except (ValueError, AttributeError):
            # Skip invalid table references
            continue
    
    return excel_tables


def get_workbook_id(file_path: str, file_name: str) -> str:
    """
    Generate a unique identifier for a workbook.
    
    Args:
        file_path: Relative path to the file
        file_name: Excel filename
        
    Returns:
        Unique workbook identifier (typically file_path/file_name)
    """
    return f"{file_path}/{file_name}"

