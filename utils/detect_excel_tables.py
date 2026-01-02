#!/usr/bin/env python3
"""
Utility script to detect Excel table objects in worksheets and generate JSONL annotations.

This script processes Excel files in a directory, checks for Excel table objects
(Worksheet.tables), and creates a JSONL file with table region information.
"""

import argparse
import json
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from utils.excel_utils import load_workbook
from openpyxl.worksheet.table import Table


def get_table_regions(worksheet) -> List[str]:
    """
    Extract table regions from a worksheet's table objects.
    
    Args:
        worksheet: openpyxl Worksheet object
        
    Returns:
        List of table region strings in format like "C8:H18"
    """
    table_regions = []
    
    # Check if worksheet has any table objects
    if hasattr(worksheet, 'tables') and worksheet.tables:
        for table_name, table_obj in worksheet.tables.items():
            # Table.ref contains the range like "C8:H18"
            if hasattr(table_obj, 'ref') and table_obj.ref:
                table_regions.append(table_obj.ref)
    
    return table_regions


def process_excel_file(file_path: Path) -> tuple[List[Dict[str, Any]], Optional[str]]:
    """
    Process a single Excel file and extract table information from all worksheets.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Tuple of (results list, error_message). error_message is None if successful.
    """
    results = []
    absolute_path = str(file_path.resolve())
    
    try:
        # Load workbook - may raise exception if file is corrupted
        workbook = load_workbook(file_path, data_only=False, read_only=True)
        
        # Iterate through all worksheets
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Check for table objects
            table_regions = get_table_regions(worksheet)
            
            # Only create entry if there are tables
            if table_regions:
                entry = {
                    "file_name": absolute_path,
                    "sheet_name": sheet_name,
                    "table_regions": table_regions
                }
                results.append(entry)
        
        workbook.close()
        return results, None
        
    except Exception as e:
        # Return error message instead of raising
        error_msg = f"{absolute_path}: {type(e).__name__}: {str(e)}"
        return results, error_msg


def process_directory(input_dir: Path, output_dir: Path, limit: int = None):
    """
    Process all Excel files in the input directory and write results to JSONL file.
    Skips corrupted files and logs errors to a timestamped error log file.
    
    Args:
        input_dir: Directory containing Excel files
        output_dir: Directory where output JSONL file will be written
        limit: Optional limit on number of files to process (for testing)
    """
    # Get all .xlsx files, sorted alphabetically
    excel_files = sorted(input_dir.glob("*.xlsx"))
    
    if limit:
        excel_files = excel_files[:limit]
    
    if not excel_files:
        print(f"No Excel files found in {input_dir}")
        return
    
    # Create output directory if it doesn't exist
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Output file path
    output_file = output_dir / "annotations_tables.jsonl"
    
    # Create error log file with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    error_log_file = output_dir / f"errors_{timestamp}.log"
    
    # Process files and write to JSONL
    # Skip corrupted files and log errors
    total_entries = 0
    error_count = 0
    
    with open(output_file, 'w', encoding='utf-8') as jsonl_f, \
         open(error_log_file, 'w', encoding='utf-8') as error_f:
        
        for excel_file in excel_files:
            print(f"Processing: {excel_file.name}", flush=True)
            entries, error_msg = process_excel_file(excel_file)
            
            if error_msg:
                # Log error and continue
                error_f.write(error_msg + '\n')
                error_count += 1
                print(f"  ERROR: Skipping corrupted file", flush=True)
            else:
                # Write successful entries
                for entry in entries:
                    json_line = json.dumps(entry, ensure_ascii=False)
                    jsonl_f.write(json_line + '\n')
                    total_entries += 1
    
    print(f"\nProcessing complete!")
    print(f"Processed {len(excel_files)} files")
    print(f"Created {total_entries} entries in {output_file}")
    if error_count > 0:
        print(f"Encountered {error_count} errors (logged to {error_log_file})")
    else:
        # Remove empty error log file if no errors
        error_log_file.unlink(missing_ok=True)


def main():
    """Main entry point for the script."""
    parser = argparse.ArgumentParser(
        description="Detect Excel table objects and generate JSONL annotations"
    )
    parser.add_argument(
        "input_dir",
        type=str,
        help="Input directory containing Excel files"
    )
    parser.add_argument(
        "output_dir",
        type=str,
        help="Output directory for JSONL file"
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limit number of files to process (for testing)"
    )
    
    args = parser.parse_args()
    
    input_path = Path(args.input_dir)
    output_path = Path(args.output_dir)
    
    if not input_path.exists():
        print(f"Error: Input directory does not exist: {input_path}")
        return 1
    
    if not input_path.is_dir():
        print(f"Error: Input path is not a directory: {input_path}")
        return 1
    
    process_directory(input_path, output_path, limit=args.limit)
    return 0


if __name__ == "__main__":
    exit(main())

