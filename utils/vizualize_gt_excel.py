#!/usr/bin/env python3
"""
Utility script to visualize ground truth table regions in Excel files.

This script reads a list of Excel file paths and an annotation JSONL file,
then marks all ground truth table regions with green background fill.
"""

import argparse
import json
import shutil
from pathlib import Path
from typing import List, Dict, Any, Optional

from utils.excel_utils import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter, column_index_from_string


# Green fill color for ground truth table regions
# Use proper Color object for Excel compatibility
GREEN_COLOR = Color(rgb="00FF00")
GREEN_FILL = PatternFill(start_color=GREEN_COLOR, end_color=GREEN_COLOR, fill_type="solid")


def parse_excel_range(range_str: str) -> tuple[tuple[int, int], tuple[int, int]]:
    """
    Parse an Excel range string (e.g., "C8:H18") into cell coordinates.
    
    Args:
        range_str: Excel range string like "C8:H18"
        
    Returns:
        Tuple of ((min_row, min_col), (max_row, max_col)) in 1-based indexing
    """
    if ':' not in range_str:
        # Single cell range
        cell = range_str
        col_letter = ''.join(c for c in cell if c.isalpha())
        row_num = int(''.join(c for c in cell if c.isdigit()))
        col_num = column_index_from_string(col_letter)
        return ((row_num, col_num), (row_num, col_num))
    
    # Range with colon
    start_cell, end_cell = range_str.split(':')
    
    # Parse start cell
    start_col_letter = ''.join(c for c in start_cell if c.isalpha())
    start_row = int(''.join(c for c in start_cell if c.isdigit()))
    start_col = column_index_from_string(start_col_letter)
    
    # Parse end cell
    end_col_letter = ''.join(c for c in end_cell if c.isalpha())
    end_row = int(''.join(c for c in end_cell if c.isdigit()))
    end_col = column_index_from_string(end_col_letter)
    
    return ((start_row, start_col), (end_row, end_col))


def mark_table_region(worksheet, range_str: str):
    """
    Mark all cells in a table region with green background fill.
    
    Args:
        worksheet: openpyxl Worksheet object
        range_str: Excel range string like "C8:H18"
    """
    try:
        (min_row, min_col), (max_row, max_col) = parse_excel_range(range_str)
        
        # Mark all cells in the range
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                # Set green fill - openpyxl will handle the cell properly
                cell.fill = GREEN_FILL
    except Exception as e:
        print(f"  Warning: Failed to mark range '{range_str}': {e}")


def load_annotations(annotation_file: Path) -> List[Dict[str, Any]]:
    """
    Load annotations from a JSONL file.
    
    Args:
        annotation_file: Path to the JSONL annotation file
        
    Returns:
        List of annotation dictionaries
    """
    annotations = []
    
    if not annotation_file.exists():
        raise FileNotFoundError(f"Annotation file not found: {annotation_file}")
    
    with open(annotation_file, 'r', encoding='utf-8') as f:
        for line_num, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            
            try:
                annotation = json.loads(line)
                annotations.append(annotation)
            except json.JSONDecodeError as e:
                print(f"Warning: Failed to parse line {line_num} in {annotation_file}: {e}")
                continue
    
    return annotations


def find_matching_annotations(annotations: List[Dict[str, Any]], excel_file_path: Path) -> List[Dict[str, Any]]:
    """
    Find all annotations that match a given Excel file.
    
    Matches by comparing the basename of the Excel file path with the
    'file_name' field in the annotations. Supports both exact matches and
    cases where the annotation file_name is a suffix of the actual filename.
    
    Args:
        annotations: List of annotation dictionaries
        excel_file_path: Absolute path to the Excel file
        
    Returns:
        List of matching annotation dictionaries
    """
    # Extract basename from the absolute path
    excel_filename = excel_file_path.name
    matching = []
    
    for annotation in annotations:
        # Compare basename to file_name attribute (which is just the filename)
        annotation_file_name = annotation.get('file_name', '')
        
        # Exact match
        if annotation_file_name == excel_filename:
            matching.append(annotation)
        # Check if annotation file_name is a suffix of the actual filename
        # (handles cases like "1_EGM - COMPETITOR GRAPHS.xlsx" matching 
        #  "VEnron2_VEnron2_1157_1_EGM - COMPETITOR GRAPHS.xlsx")
        elif excel_filename.endswith(annotation_file_name):
            matching.append(annotation)
        # Also check reverse (actual filename is suffix of annotation)
        elif annotation_file_name.endswith(excel_filename):
            matching.append(annotation)
    
    return matching


def visualize_gt_for_file(
    excel_file_path: Path,
    annotations: List[Dict[str, Any]],
    output_dir: Optional[Path] = None
) -> bool:
    """
    Visualize ground truth table regions for a single Excel file.
    
    Args:
        excel_file_path: Absolute path to the Excel file to process
        annotations: List of all annotations
        output_dir: Optional output directory. If None, saves next to input file.
        
    Returns:
        True if successful, False otherwise
    """
    if not excel_file_path.exists():
        print(f"Error: Excel file not found: {excel_file_path}")
        return False
    
    # Find matching annotations for this file
    matching_annotations = find_matching_annotations(annotations, excel_file_path)
    
    if not matching_annotations:
        print(f"  No annotations found for {excel_file_path.name}")
        return False
    
    # Determine output path first (before copying)
    if output_dir:
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / excel_file_path.name.replace('.xlsx', '_viz.xlsx')
    else:
        # Save next to input file
        output_path = excel_file_path.parent / excel_file_path.name.replace('.xlsx', '_viz.xlsx')
    
    # Copy the original file to the output path (preserves all metadata and structure)
    try:
        shutil.copy2(excel_file_path, output_path)
    except Exception as copy_error:
        print(f"  Error copying file to {output_path}: {copy_error}")
        return False
    
    try:
        # Load workbook from the copied file (not read-only, since we need to modify it)
        # Try loading normally first to preserve all workbook properties
        try:
            workbook = load_workbook(output_path, data_only=False)
        except Exception as e:
            # If that fails due to XML/parsing errors, try with keep_links=False
            error_msg = str(e).lower()
            if "parseerror" in error_msg or "not well-formed" in error_msg or "invalid token" in error_msg:
                try:
                    workbook = load_workbook(output_path, data_only=False, keep_links=False)
                except Exception:
                    # If that also fails, try without VBA
                    workbook = load_workbook(output_path, data_only=False, keep_links=False, keep_vba=False)
            else:
                # Re-raise the original error if it's not a parsing error
                raise e
        
        # Process each matching annotation
        sheets_processed = 0
        for annotation in matching_annotations:
            sheet_name = annotation.get('sheet_name')
            table_regions = annotation.get('table_regions', [])
            
            if not sheet_name:
                print(f"  Warning: Missing 'sheet_name' in annotation, skipping")
                continue
            
            if not table_regions:
                print(f"  Warning: No 'table_regions' in annotation for sheet '{sheet_name}', skipping")
                continue
            
            # Check if sheet exists
            if sheet_name not in workbook.sheetnames:
                print(f"  Warning: Sheet '{sheet_name}' not found in workbook, skipping")
                continue
            
            worksheet = workbook[sheet_name]
            
            # Mark each table region
            for region in table_regions:
                mark_table_region(worksheet, region)
            
            sheets_processed += 1
            print(f"  Processed sheet '{sheet_name}': {len(table_regions)} table region(s)")
        
        if sheets_processed == 0:
            print(f"  No sheets were processed for {excel_file_path.name}")
            workbook.close()
            # Clean up the copied file if no sheets were processed
            try:
                output_path.unlink()
            except:
                pass
            return False
        
        # Save the workbook back to the copied file (overwrite in-place)
        # Ensure workbook is not in write_only mode
        if hasattr(workbook, 'write_only') and workbook.write_only:
            print(f"  Warning: Workbook is in write_only mode, this may cause issues")
        
        # Save with explicit options to ensure compatibility
        # Use a try-except to handle any save errors gracefully
        try:
            # Save the workbook back to the copied file
            workbook.save(output_path)
        except Exception as save_error:
            try:
                workbook.close()
            except:
                pass
            print(f"  Error saving workbook: {save_error}")
            import traceback
            traceback.print_exc()
            # Clean up the copied file if save failed
            try:
                output_path.unlink()
            except:
                pass
            return False
        
        # Close the workbook after saving
        workbook.close()
        
        # Verify the file was saved correctly by trying to open it
        try:
            verify_wb = load_workbook(output_path, read_only=True)
            verify_wb.close()
        except Exception as verify_error:
            print(f"  Warning: Saved file may be corrupted - verification failed: {verify_error}")
            # Don't return False here - file might still be valid for Excel even if openpyxl can't verify
            pass
        
        print(f"  Saved visualization to: {output_path}")
        return True
        
    except Exception as e:
        error_type = type(e).__name__
        error_msg = str(e)
        
        # Provide more helpful error messages for common issues
        if "ParseError" in error_type or "not well-formed" in error_msg:
            print(f"  Error processing {excel_file_path.name}: XML parsing error - file may be corrupted or have encoding issues")
            print(f"    Details: {error_msg[:100]}...")
        else:
            print(f"  Error processing {excel_file_path.name}: {error_type}: {error_msg}")
        return False


def process_file_list(
    file_list_path: Path,
    annotation_file: Path,
    output_dir: Optional[Path] = None
):
    """
    Process a list of Excel files and visualize their ground truth table regions.
    
    Args:
        file_list_path: Path to file containing absolute Excel file paths (one per line)
        annotation_file: Path to JSONL annotation file
        output_dir: Optional output directory. If None, saves next to each input file.
    """
    # Load annotations
    print(f"Loading annotations from: {annotation_file}")
    annotations = load_annotations(annotation_file)
    print(f"Loaded {len(annotations)} annotation entries")
    
    # Read Excel file paths (absolute paths)
    if not file_list_path.exists():
        print(f"Error: File list not found: {file_list_path}")
        return
    
    excel_file_paths = []
    with open(file_list_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#'):  # Skip empty lines and comments
                file_path = Path(line)
                # Skip temp files (Excel creates these when files are open)
                if file_path.name.startswith('~$'):
                    print(f"Skipping temp file: {file_path.name}")
                    continue
                excel_file_paths.append(file_path)
    
    if not excel_file_paths:
        print(f"No Excel file paths found in {file_list_path}")
        return
    
    print(f"\nProcessing {len(excel_file_paths)} Excel file(s)...")
    
    # Process each Excel file
    success_count = 0
    for excel_file_path in excel_file_paths:
        print(f"\nProcessing: {excel_file_path}")
        
        if visualize_gt_for_file(excel_file_path, annotations, output_dir):
            success_count += 1
    
    print(f"\n{'='*60}")
    print(f"Processing complete!")
    print(f"Successfully processed: {success_count}/{len(excel_file_paths)} files")


def main():
    """Main entry point for the script."""
    parser = argparse.ArgumentParser(
        description="Visualize ground truth table regions in Excel files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process files and save visualizations next to input files
  python utils/vizualize_gt_excel.py file_list.txt annotations.jsonl
  
  # Process files and save to a specific output directory
  python utils/vizualize_gt_excel.py file_list.txt annotations.jsonl -o output_dir/
        """
    )
    
    parser.add_argument(
        "file_list",
        type=str,
        help="File containing absolute Excel file paths (one per line)"
    )
    
    parser.add_argument(
        "annotation_file",
        type=str,
        help="Path to JSONL annotation file"
    )
    
    parser.add_argument(
        "-o", "--output-dir",
        type=str,
        default=None,
        help="Output directory for visualization files. If not specified, "
             "files are saved next to input files with '_viz.xlsx' suffix."
    )
    
    args = parser.parse_args()
    
    file_list_path = Path(args.file_list)
    annotation_file_path = Path(args.annotation_file)
    output_dir = Path(args.output_dir) if args.output_dir else None
    
    if not file_list_path.exists():
        print(f"Error: File list does not exist: {file_list_path}")
        return 1
    
    if not annotation_file_path.exists():
        print(f"Error: Annotation file does not exist: {annotation_file_path}")
        return 1
    
    process_file_list(file_list_path, annotation_file_path, output_dir)
    return 0


if __name__ == "__main__":
    exit(main())

